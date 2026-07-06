import os
import json
import requests
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

load_dotenv()
TENANT_ID = os.getenv("AZURE_TENANT_ID")
CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
LEGACY_AUTH_CLIENT_TYPES = {"exchangeActiveSync", "other"}

if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET]):
    raise SystemExit(
        "Missing credentials. Copy .env.example to .env and fill in "
        "AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET."
    )

def get_access_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    resp = requests.post(url, data=data)
    resp.raise_for_status()
    return resp.json()["access_token"]

def graph_get_all(url, headers):
    results = []
    while url:
        resp = requests.get(url, headers=headers)
        if resp.status_code >= 400:
            print("ERROR RESPONSE BODY:", resp.text)
        resp.raise_for_status()
        body = resp.json()
        results.extend(body.get("value", []))
        url = body.get("@odata.nextLink")
    return results

def load_known_accounts():
    path = "known_accounts.json"
    if not os.path.exists(path):
        print(f"[warn] {path} not found — break-glass/service accounts won't be labeled. "
              f"Copy known_accounts.json.example to get started.")
        return {"break_glass_accounts": [], "service_accounts": []}
    with open(path) as f:
        data = json.load(f)
    return {
        "break_glass_accounts": {u.lower() for u in data.get("break_glass_accounts", [])},
        "service_accounts": {u.lower() for u in data.get("service_accounts", [])},
    }

def fetch_mfa_registration(headers):
    url = f"{GRAPH_BASE}/reports/authenticationMethods/userRegistrationDetails"
    return graph_get_all(url, headers)

def fetch_ca_policies(headers):
    url = f"{GRAPH_BASE}/identity/conditionalAccess/policies"
    return graph_get_all(url, headers)

def policy_requires_mfa(policy):
    controls = policy.get("grantControls") or {}
    return "mfa" in (controls.get("builtInControls") or [])
 
 
def policy_blocks_legacy_auth(policy):
    client_types = set((policy.get("conditions") or {}).get("clientAppTypes") or [])
    grant = policy.get("grantControls") or {}
    is_block = grant.get("operator") == "OR" and "block" in (grant.get("builtInControls") or [])
    return is_block and bool(client_types & LEGACY_AUTH_CLIENT_TYPES)
 
 
def user_in_policy_scope(user_id, policy):
    """Rough scope check: included (all users, or explicit id) minus excluded."""
    users_cond = (policy.get("conditions") or {}).get("users") or {}
    include_users = set(users_cond.get("includeUsers") or [])
    exclude_users = set(users_cond.get("excludeUsers") or [])
 
    if user_id in exclude_users:
        return False
    if "All" in include_users or user_id in include_users:
        return True
    # Note: group-based include/exclude requires resolving group membership,
    # which needs GroupMember.Read.All. Left as a documented gap below.
    return False

def build_report(mfa_records, policies, known):
    enabled_mfa_policies = [
        p for p in policies if p.get("state") == "enabled" and policy_requires_mfa(p)
    ]
    reportonly_mfa_policies = [
        p for p in policies
        if p.get("state") == "enabledForReportingButNotEnforced" and policy_requires_mfa(p)
    ]
    legacy_auth_open = not any(
        p.get("state") == "enabled" and policy_blocks_legacy_auth(p) for p in policies
    )
 
    rows = []
    for u in mfa_records:
        user_id = u.get("id")
        upn = (u.get("userPrincipalName") or "").lower()
        is_admin = u.get("isAdmin", False)
        mfa_registered = u.get("isMfaRegistered", False)
 
        covered_enforced = any(user_in_policy_scope(user_id, p) for p in enabled_mfa_policies)
        covered_reportonly = any(user_in_policy_scope(user_id, p) for p in reportonly_mfa_policies)
 
        label = None
        if upn in known["break_glass_accounts"]:
            label = "break_glass"
        elif upn in known["service_accounts"]:
            label = "service_account"
 
        # Determine flag
        if label == "break_glass":
            flag = "INFO"
            reason = "Known break-glass account — expected to bypass MFA/CA by design."
        elif not mfa_registered and not covered_enforced:
            flag = "RED"
            reason = "No MFA registered AND not covered by any enforced CA policy requiring MFA."
        elif not covered_enforced and covered_reportonly:
            flag = "YELLOW"
            reason = "Only covered by a report-only CA policy — MFA not actually enforced yet."
        elif not covered_enforced:
            flag = "YELLOW"
            reason = "MFA registered but no enforced CA policy actually requires it for this user."
        elif not mfa_registered:
            flag = "YELLOW"
            reason = "Covered by an enforcing CA policy, but user hasn't completed MFA registration."
        else:
            flag = "GREEN"
            reason = "MFA registered and enforced via Conditional Access."
 
        rows.append({
            "userPrincipalName": u.get("userPrincipalName"),
            "isAdmin": is_admin,
            "accountLabel": label or ("unclassified" if flag != "GREEN" else ""),
            "mfaRegistered": mfa_registered,
            "coveredByEnforcedCA": covered_enforced,
            "coveredByReportOnlyCA": covered_reportonly,
            "flag": flag,
            "reason": reason,
        })
 
    return rows, legacy_auth_open

def rank_remediation(rows, legacy_auth_open):
    """Produce a prioritized action list."""
    items = []
 
    admins_no_mfa = [r for r in rows if r["isAdmin"] and r["flag"] == "RED"]
    if admins_no_mfa:
        items.append({
            "priority": 1,
            "category": "Privileged accounts with no protection",
            "count": len(admins_no_mfa),
            "accounts": [r["userPrincipalName"] for r in admins_no_mfa],
            "action": "Enforce MFA immediately via Conditional Access for these admin accounts.",
        })
 
    zero_coverage = [r for r in rows if r["flag"] == "RED" and not r["isAdmin"]]
    if zero_coverage:
        items.append({
            "priority": 2,
            "category": "Users with zero CA/MFA coverage",
            "count": len(zero_coverage),
            "accounts": [r["userPrincipalName"] for r in zero_coverage],
            "action": "Add these users to an enforced CA policy requiring MFA.",
        })
 
    if legacy_auth_open:
        items.append({
            "priority": 3,
            "category": "Legacy authentication not blocked",
            "count": None,
            "accounts": [],
            "action": "No enabled CA policy blocks legacy auth protocols (exchangeActiveSync/other). "
                       "Create/enable a policy to block legacy auth tenant-wide.",
        })
 
    report_only = [r for r in rows if r["flag"] == "YELLOW" and r["coveredByReportOnlyCA"]]
    if report_only:
        items.append({
            "priority": 4,
            "category": "Policies in report-only mode",
            "count": len(report_only),
            "accounts": [r["userPrincipalName"] for r in report_only],
            "action": "Review report-only CA policies and switch to Enabled once validated.",
        })
 
    unclassified = [r for r in rows if r["accountLabel"] == "unclassified"]
    if unclassified:
        items.append({
            "priority": 5,
            "category": "Unclassified excluded/unprotected accounts",
            "count": len(unclassified),
            "accounts": [r["userPrincipalName"] for r in unclassified],
            "action": "Verify whether these are legitimate service/break-glass accounts. "
                       "Add confirmed ones to known_accounts.json.",
        })
 
    return items
 
 
def write_xlsx(rows, remediation, out_path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Coverage Heatmap"
    headers = ["User", "Admin?", "Label", "MFA Registered", "Enforced CA Coverage",
               "Report-Only CA Coverage", "Flag", "Reason"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    fill = {
        "RED": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "YELLOW": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "GREEN": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "INFO": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    for r in sorted(rows, key=lambda x: {"RED": 0, "YELLOW": 1, "INFO": 2, "GREEN": 3}[x["flag"]]):
        ws1.append([
            r["userPrincipalName"], r["isAdmin"], r["accountLabel"], r["mfaRegistered"],
            r["coveredByEnforcedCA"], r["coveredByReportOnlyCA"], r["flag"], r["reason"],
        ])
        ws1.cell(row=ws1.max_row, column=7).fill = fill[r["flag"]]

    for col in ws1.columns:
        length = max(len(str(c.value)) for c in col if c.value is not None)
        ws1.column_dimensions[col[0].column_letter].width = min(length + 2, 50)

    ws2 = wb.create_sheet("Remediation List")
    ws2.append(["Priority", "Category", "Count", "Action", "Example Accounts (see Full Account Lists tab)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for item in remediation:
        ws2.append([
            item["priority"], item["category"], item["count"] or "-",
            item["action"], ", ".join(item["accounts"][:5]) + ("..." if len(item["accounts"]) > 5 else ""),
        ])
    for col in ws2.columns:
        length = max(len(str(c.value)) for c in col if c.value is not None)
        ws2.column_dimensions[col[0].column_letter].width = min(length + 2, 60)

    ws3 = wb.create_sheet("Full Account Lists")
    ws3.append(["Priority", "Category", "Account"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for item in remediation:
        for account in item["accounts"]:
            ws3.append([item["priority"], item["category"], account])
    for col in ws3.columns:
        length = max(len(str(c.value)) for c in col if c.value is not None)
        ws3.column_dimensions[col[0].column_letter].width = min(length + 2, 50)

    wb.save(out_path)


def main():
    print("Authenticating...")
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
 
    print("Fetching MFA registration details...")
    mfa_records = fetch_mfa_registration(headers)
    print(f"  -> {len(mfa_records)} users")
 
    print("Fetching Conditional Access policies...")
    policies = fetch_ca_policies(headers)
    print(f"  -> {len(policies)} policies")
 
    known = load_known_accounts()
 
    print("Building coverage report...")
    rows, legacy_auth_open = build_report(mfa_records, policies, known)
    remediation = rank_remediation(rows, legacy_auth_open)
 
    os.makedirs("output", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"output/mfa_ca_coverage_report_{timestamp}.xlsx"
    write_xlsx(rows, remediation, out_path)
 
    red = sum(1 for r in rows if r["flag"] == "RED")
    yellow = sum(1 for r in rows if r["flag"] == "YELLOW")
    green = sum(1 for r in rows if r["flag"] == "GREEN")
    print(f"\nDone. Red: {red}  Yellow: {yellow}  Green: {green}")
    print(f"Report saved to {out_path}")
 
 
if __name__ == "__main__":
    main()